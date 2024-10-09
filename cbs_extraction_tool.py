from sqlalchemy import create_engine, text,inspect
from sqlalchemy.orm import sessionmaker
import re
import pandas as pd
import os
import subprocess
import json
import numpy as np
import pymysql.err
import io
import shutil
from datetime import datetime, timedelta
from dateutil import parser
from openpyxl import load_workbook


# DEFINE THE DATABASE CREDENTIALS
user = 'root'
password = 'root'
host = '127.0.0.1'
port = 3306
 
# PYTHON FUNCTION TO CONNECT TO THE MYSQL DATABASE AND
# RETURN THE SQLACHEMY ENGINE OBJECT
def get_connection():
    return create_engine(
        url="mysql+pymysql://{0}:{1}@{2}:{3}".format(
            user, password, host, port
        )
    )

def get_db_connection(database):
    return create_engine(
        url="mysql+pymysql://{0}:{1}@{2}:{3}/{4}".format(
            user, password, host, port, database
        )
    )

def get_all_files(root_dir):
    all_files = []
    for root, directories, files in os.walk(root_dir):
        for file in files:
            all_files.append(os.path.join(root, file))
    all_files = [path.replace("\\","/").replace("._", "") for path in all_files]
    return all_files

def calculate_age_at_visit(row):

    try:
        birthdate = pd.to_datetime(row["contact_birthdate"], errors='coerce')
        date_of_hiv_test = pd.to_datetime(row["time_contact_tested"], errors='coerce')
        if date_of_hiv_test is None:
            return -1
        # Convert the dates to datetime.datetime objects before subtracting them.
        birthdate_dt = birthdate.to_pydatetime()
        date_of_hiv_test_dt = date_of_hiv_test.to_pydatetime()
        
        # Calculate the age at visit.
        age_at_visit = (date_of_hiv_test_dt - birthdate_dt).days / 365.25
        age_at_visit = np.floor(age_at_visit)
    except TypeError:
        age_at_visit = -1
    except Exception as e:
        age_at_visit = -1
    except pd.errors.OutOfBoundsDatetime as err:
        # Log the error and continue to the next iteration of the loop
        age_at_visit = -1
    return age_at_visit

def get_site_id(metadata):
  """Returns the siteId from the metadata string."""
  match = re.search(r"<string>siteId</string><string>(.*?)</string>", metadata)
  if match:
    return match.group(1)
  else:
    return None

def get_time(metadata):
  """Returns the time from the metadata string."""
  match = re.search(r"<string>time</string><string>(.*?)</string>", metadata)
  if match:
    return match.group(1)
  else:
    return None

def get_version(metadata):
  """Returns the version from the metadata string."""
  match = re.search(r"<string>version</string><string>(.*?)</string>", metadata)
  if match:
    return match.group(1)
  else:

    return None
def get_username(metadata):
  """Returns the username from the metadata string."""
  match = re.search(r"<string>username</string><string>(.*?)</string>", metadata)
  if match:
    return match.group(1)
  else:
    return None

def run_script(script_path):
  subprocess.run(["python", script_path], timeout=600)

def calculate_last_visit_date(df):
    df['art_visit_date'] = pd.to_datetime(df['art_visit_date'], errors="coerce")
    df.sort_values(by=['person_id', 'art_visit_date'], inplace=True)
    df['last_art_visit_date'] = df.groupby('person_id')['art_visit_date'].transform("max")
    return df

class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime.datetime):
            return obj.strftime("%Y-%m-%d")
        else:
            return super().default(obj)
        

def append_df_to_excel(filename, df, sheet_name='Sheet1', **to_excel_kwargs):
    # Check if the file exists
    if os.path.exists(filename):
        # File exists, open in append mode
        mode = 'a'
        # Load the workbook to check the last sheet's row count
        book = load_workbook(filename)
        # Get the last sheet name
        last_sheet_name = book.sheetnames[-1]
        # Check if the last sheet has reached the row limit
        if book[last_sheet_name].max_row >= 1048576:
            # Create a new sheet name
            sheet_name = 'Sheet' + str(len(book.sheetnames) + 1)
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode=mode, if_sheet_exists='overlay')
    else:
        # File does not exist, create a new file
        mode = 'w'
        # No need to load the workbook or check the row count
    
        # Create a Pandas Excel writer using openpyxl as the engine
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode=mode)

    # Convert the dataframe to an XlsxWriter Excel object
    df.to_excel(writer, sheet_name=sheet_name, index=False, **to_excel_kwargs)

    # Close the Pandas Excel writer and output the Excel file
    writer.close()
 

if __name__ == '__main__':  
    index_data = []
    

    backup_folder =  "/home/devoop/Documents/eHR/mutare"
    backup_files = get_all_files(backup_folder)
   

    # Restore each backup file
    for backup_file in set(backup_files): 
        try:
             # GET THE CONNECTION OBJECT (ENGINE) FOR THE DATABASE
            engine = get_connection()
            # Create a session
            Session = sessionmaker(bind=engine)
            session = Session()
            print( f"Connection to the {host} for user {user} created successfully.")
            with open("./processed.txt", "r")  as processed_log_file:
                    extracted_sites = [file for file in processed_log_file]
                    extracted_sites = [i.replace("\n", "") for i in extracted_sites]
            if backup_file.endswith(".sql") and backup_file not in extracted_sites and backup_file not in["/home/devoop/Documents/eHR/mutare"]:
                print(f"{backup_file} NOT YET PROCESSED...")
                facility_art_data = {}
                with engine.connect() as connection:
                    results = connection.execute(text('SELECT schema_name FROM information_schema.schemata;'))
                    
                    schemas = [schema[0] for schema in results]
                    for schema in schemas:
                        if schema in ['client', 'consultation', 'deduplication', 'facility', 'mrs', 'provider', 'report', 'terminology', 'zimepms']:
                            connection.execute(text(f'SET foreign_key_checks = 0'))

                            results = connection.execute(text(f'DROP DATABASE {schema}'))
                            print(f"DROPPED [{schema}]")
                    connection.close()

                restore_command = f"""mysql -h 127.0.0.1 -u root -proot -f --binary-mode=1 < {backup_file}"""
                print(f'RESTORING DATABASE FOR [{backup_file}]')
                try:
                    subprocess.run(restore_command, shell=True, check=True)
                    # with open(backup_file, 'r') as file:
                    #     restore_command = f"mysql -u root -ppassword -f --binary-mode=1}"
                    #     subprocess.run(restore_command, shell=True, input=file.read(), check=True)
                except Exception as e:
                    with open('./logs.txt', 'a') as f:
                        f.write(f'{backup_file}\n')
                    continue
                finally:
                    engine.dispose()

                mrs_database = 'mrs'
                print( f"Connecting to the {mrs_database} database for {backup_file}")
                mrs_engine = get_db_connection(mrs_database)
                print( f"Connection successful")
                print(f"Getting site code...")
                mrs_df = pd.read_sql(f'SELECT meta_data FROM domain_event_entry WHERE time_stamp = (SELECT MAX(time_stamp) FROM domain_event_entry)', con=mrs_engine)
                mrs_engine.dispose()
                if len(mrs_df['meta_data']) != 0:
                    meta_data = mrs_df['meta_data'].iloc[0].decode("utf-8")
                    site_id = get_site_id(meta_data)
                    time = get_time(meta_data)
                    version = get_version(meta_data)
                    username = get_username(meta_data)
                    print(f"*** Fetching site code successful *** SITE_CODE: [{site_id}] - EHR_VERSION: [{version}] - LAST_USED [{time}]")
                    facility_mapping_file = pd.read_csv("./Mapping_file_05_feb_2024_v3.csv")
                    province_dict = dict(zip(facility_mapping_file["Facility ID"], facility_mapping_file["Province"]))
                    district_dict = dict(zip(facility_mapping_file["Facility ID"], facility_mapping_file["District"]))
                    faclity_dict = dict(zip(facility_mapping_file["Facility ID"], facility_mapping_file["Facility"]))
                    province = province_dict.get(site_id)
                    district = district_dict.get(site_id)
                    facility = faclity_dict.get(site_id)
                    print(f"**** {province} - {district} - {facility} ****")

                    # SQL query to join the tables and filter the results
                    positive_patients_query = """
                    SELECT p.person_id, h.laboratory_investigation_id, h.hts_number, h.purpose AS reason_for_hiv_testing, h.time AS date_of_hiv_test, h.entry_point, h.pregnant AS pregnant_during_hiv_test, h.client_already_positive, h.client_already_on_art, h.hts_type AS test_method, pi.result AS hiv_final_test_result, h.reason_for_not_performing_test AS reason_for_not_performing_recency_test
                    FROM hts h
                    JOIN patient p ON h.patient_id = p.patient_id
                    JOIN person_investigation pi ON p.person_id = pi.person_id
                    WHERE pi.test = 'HIV' AND LOWER(pi.result) = 'positive'
                    UNION ALL
                    SELECT c.person_id, NULL AS laboratory_investigation_id, NULL AS hts_number, NULL AS reason_for_hiv_testing, NULL AS date_of_hiv_test, NULL AS entry_point, NULL AS pregnant_during_hiv_test, NULL AS client_already_positive, NULL AS client_already_on_art, NULL AS test_method, NULL AS hiv_final_test_result, NULL AS reason_for_not_performing_recency_test
                    FROM cbs c
                    """
                    
                    print(f'FETCHING POSITIVE CLIENTS...')
                    # Execute the query and fetch the result into a DataFrame
                    engine = get_db_connection("consultation")
                    pos_df = pd.read_sql(positive_patients_query, engine)
                    engine.dispose()

                    if len(pos_df) != 0:
                        print(f'FETCHING PATIENT IDs...')
                        # SQL query to join the tables and filter the results
                        patient_id_query = """
                        SELECT person_id, patient_id FROM patient
                        """

                        # Execute the query and fetch the result into a DataFrame
                        engine = get_db_connection("consultation")
                        patient_id_df = pd.read_sql(patient_id_query, engine)
                        engine.dispose()

                        pos_df = pos_df.merge(patient_id_df, on="person_id")

                        print(f'FETCHING DEMOGRAPHICS...')
                        # Extract the list of person_ids
                        pos_person_ids = list(pos_df['person_id'].unique())

                        # SQL query to select the specified details for the person_ids
                        query = f"""
                        SELECT person_id, town, city, street, birthdate, sex, self_identified_gender, marital, education, occupation, religion, nationality
                        FROM person
                        WHERE person_id IN {tuple(pos_person_ids)}
                        """
                        if len(pos_person_ids) > 2:
                            # Execute the query and fetch the result into a DataFrame
                            engine = get_db_connection("client")
                            pos_person_demos = pd.read_sql(query, engine)
                            engine.dispose()

                            hts_with_demos = pos_df.merge(pos_person_demos, on="person_id")
                        elif len(pos_person_ids) >= 1 and  len(pos_person_ids) <= 2:
                            
                            pos_person_ids_2 = ', '.join(f"'{uuid}'" for uuid in pos_person_ids)
                            query = f"""
                                SELECT person_id, town, city, street, birthdate, sex, self_identified_gender, marital, education, occupation, religion, nationality
                                FROM person
                                WHERE person_id IN ({str(pos_person_ids_2).replace(",","")})
                                """
                             # Execute the query and fetch the result into a DataFrame
                            engine = get_db_connection("client")
                            pos_person_demos = pd.read_sql(query, engine)
                            engine.dispose()
                            hts_with_demos = pos_df.merge(pos_person_demos, on="person_id")

                        # Query the laboratory_investigation_test table for each laboratory_investigation_id
                      
                        for index, row in hts_with_demos.iterrows():
                            laboratory_investigation_id = row['laboratory_investigation_id']
                            test_method = row['test_method']
                            
                            query = f"""
                            SELECT * FROM laboratory_investigation_test
                            WHERE laboratory_investigation_id = '{laboratory_investigation_id}'
                            ORDER BY time;
                            """
                            engine = get_db_connection("consultation")
                            df_test_results = pd.read_sql(query, engine)
                            engine.dispose()
                            
                            # Check if df_test_results is not empty and contains at least two rows
                            if not df_test_results.empty and len(df_test_results) >= 1:
                                # Extractdf test_results based on test_method
                                if test_method == "SELF_TEST":
                                    hts_with_demos.at[index, 'hiv_test_one_result'] = df_test_results.iloc[0]['result']
                                if len(df_test_results) == 1:
                                    hts_with_demos.at[index, 'hiv_test_one_result'] = df_test_results.iloc[0]['result']
                                elif len(df_test_results) == 2:
                                    hts_with_demos.at[index, 'hiv_test_one_result'] = df_test_results.iloc[0]['result']
                                    hts_with_demos.at[index, 'hiv_test_two_result'] = df_test_results.iloc[1]['result']
                                elif len(df_test_results) > 2:
                                  
                                    hts_with_demos.at[index, 'hiv_test_one_result'] = df_test_results.iloc[0]['result']
                                    hts_with_demos.at[index, 'hiv_test_two_result'] = df_test_results.iloc[1]['result']
                                    hts_with_demos.at[index, 'hiv_test_three_result'] = df_test_results.iloc[2]['result']
                            else:
                                # Handle the case where df_test_results is empty or does not contain enough rows

                                hts_with_demos.at[index, 'hiv_test_one_result'] = None # or any other default value
                                hts_with_demos.at[index, 'hiv_test_two_result'] = None # or any other default value
                                hts_with_demos.at[index, 'hiv_test_three_result'] = None # or any other default value
                            
                        # For Back captured Data
                        # Filter rows with missing laboratory_investigation_id
                        rows_with_missing_id = hts_with_demos[pd.isnull(hts_with_demos['laboratory_investigation_id'])]
                     
                        person_ids_to_query = tuple(rows_with_missing_id['person_id'])

                        # Construct the query to fetch test results and date for the given person_ids
                        if len(person_ids_to_query) > 2:
                            query = f"""
                            SELECT person_id, 'POSITIVE' AS hiv_test_one_result, 'POSITIVE' AS hiv_test_two_result, 'POSITIVE' AS hiv_test_three_result, 'POSITIVE' AS hiv_final_test_result, date_of_hiv_test, reason_for_not_initiating_art
                            FROM cbs
                            WHERE person_id IN {person_ids_to_query}
                            """
                            engine = get_db_connection("consultation")
                            # Execute the query and fetch the results
                            results = pd.read_sql(query, engine)
                            engine.dispose()
                             # Set 'person_id' as the index for both DataFrames
                            hts_with_demos.set_index('person_id', inplace=True)
                            results.set_index('person_id', inplace=True)

                            # Update hts_with_demos with the values from results
                            hts_with_demos.update(results)
                            # Reset the index if you need the original index back
                            hts_with_demos.reset_index(inplace=True)
                        elif len(person_ids_to_query) >= 1 and  len(person_ids_to_query) <= 2:
                            
                            person_ids_to_query = ', '.join(f"'{uuid}'" for uuid in pos_person_ids)
                            query = f"""
                            SELECT person_id, 'POSITIVE' AS hiv_test_one_result, 'POSITIVE' AS hiv_test_two_result, 'POSITIVE' AS hiv_test_three_result, 'POSITIVE' AS hiv_final_test_result, date_of_hiv_test, reason_for_not_initiating_art
                            FROM cbs
                            WHERE person_id IN ({str(person_ids_to_query).replace(",","")})
                            """
                            engine = get_db_connection("consultation")
                            # Execute the query and fetch the results
                            results = pd.read_sql(query, engine)
                            engine.dispose()
                             # Set 'person_id' as the index for both DataFrames
                            hts_with_demos.set_index('person_id', inplace=True)
                            results.set_index('person_id', inplace=True)

                            # Update hts_with_demos with the values from results
                            hts_with_demos.update(results)
                            # Reset the index if you need the original index back
                            hts_with_demos.reset_index(inplace=True)
                       

                        # Prepare the SQL Query
                        if len(pos_person_ids) > 2:
                            query = f"""
                            SELECT person_id, date AS date_recency_done, result AS recency_result
                            FROM person_investigation
                            WHERE test='HIV-1 Rapid Recency' AND person_id IN {tuple(pos_person_ids)}
                            """
                            engine = get_db_connection("consultation")
                            # Execute the Query and Fetch Results
                            results = pd.read_sql(query, engine)
                            engine.dispose()

                            # Merge the DataFrames
                            hts_with_demos = pd.merge(hts_with_demos, results, on='person_id', how='left')
                        elif len(pos_person_ids) >= 1 and  len(pos_person_ids) <= 2:
                            person_ids_to_query = ', '.join(f"'{uuid}'" for uuid in pos_person_ids)
                            query = f"""
                            SELECT person_id, date AS date_recency_done, result AS recency_result
                            FROM person_investigation
                            WHERE test='HIV-1 Rapid Recency' AND person_id IN ({str(person_ids_to_query).replace(",","")})
                            """
                            engine = get_db_connection("consultation")
                            # Execute the Query and Fetch Results
                            results = pd.read_sql(query, engine)
                            engine.dispose()

                        art_query = """ 
                                select  art_id, person_id, art_number, art_cohort_number, date AS Art_Initiation_Date, date_enrolled AS date_enrolled_into_art
                                from art 
                            """
                        engine = get_db_connection("consultation")
                        art_results = pd.read_sql(art_query, engine)
                        engine.dispose()

                        hts_art_with_demos = pd.merge(hts_with_demos, art_results, on='person_id', how='left')


                        art_current_status_query = """
                            select a.date as art_current_status_date, a.regimen , a.state as arv_status, a.art_initiation_category, art.person_id, a.regimen_id AS art_regimen_id
                            from art_current_status a 
                            left join art 
                            on a.art_id = art.art_id 
                        """     
                        engine = get_db_connection("consultation")
                        art_current_status_results = pd.read_sql(art_current_status_query, engine)
                        engine.dispose()
                        hts_art_art_current_status_with_demos = pd.merge(hts_art_with_demos, art_current_status_results, on='person_id', how='left')

                        art_who_stage_query = """
                            select a.art_id, a.date as who_stage_date, a.date AS who_outcome_date, a.stage AS who_stage, a.follow_up_status AS art_outcome                    
                            from art_who_stage a left join art p
                            on a.art_id  = p.art_id   
                        """
                        engine = get_db_connection("consultation")
                        art_current_status_who_results = pd.read_sql(art_who_stage_query, engine)
                        engine.dispose()
                        hts_art_art_current_status_who_with_demos = pd.merge(hts_art_art_current_status_with_demos, art_current_status_who_results, on='art_id', how='left')

                        viral_load_query = """
                            select person_id, date as date_at_which_viral_load_result_was_issued, date as date_for_which_viral_load_was_taken, 'TRUE' as viral_load_sample_submitted_to_lab,
                            'TRUE' as was_viral_load_result_issued,
                            result as viral_load_result
                            from person_investigation where test = 'Viral Load'      
                        """
                        engine = get_db_connection("consultation")
                        art_current_status_who_vl_results = pd.read_sql(viral_load_query, engine)
                        engine.dispose()
                        hts_art_art_current_status_who_vl_with_demos = pd.merge(hts_art_art_current_status_who_with_demos, art_current_status_who_vl_results, on='person_id', how='left')


                        cd4_query = """
                            select person_id, date AS date_at_which_cd4_sample_was_taken, date as date_at_which_cd4_result_was_issued, 'TRUE' as cd4_sample_submitted_to_lab,
                            'TRUE' as was_cd4_result_issued,
                            result as cd4_count
                            from person_investigation where test = 'CD4 Count' 
                        """  

                        engine = get_db_connection("consultation")
                        art_current_status_who_vl_cd4_results = pd.read_sql(cd4_query, engine)
                        engine.dispose()
                        hts_art_art_current_status_who_vl_cd4_with_demos = pd.merge(hts_art_art_current_status_who_vl_with_demos, art_current_status_who_vl_cd4_results, on='person_id', how='left')

                        TB_query = """
                            select person_id, date AS tb_treatment_start_date, type_of_tb, tb_disease_site, tb_disease_type, outcome as tb_outcome
                            from tb
                        """
                        engine = get_db_connection("consultation")
                        art_current_status_who_vl_cd4_tb_results = pd.read_sql(TB_query, engine)
                        engine.dispose()

                        hts_art_art_current_status_who_vl_cd4_tb_with_demos = pd.merge(hts_art_art_current_status_who_vl_cd4_with_demos, art_current_status_who_vl_cd4_tb_results, on='person_id', how='left')

                        TB_Screening_query = """
                            select t.person_id , p.presumptive AS tb_screened
                            from patient_tb_screening p
                            left join patient t 
                            on p.patient_id = t.patient_id
                        """  
                        engine = get_db_connection("consultation")
                        art_current_status_who_vl_cd4_tb_tbscreen_results = pd.read_sql(TB_Screening_query, engine)
                        engine.dispose()
                        hts_art_art_current_status_who_vl_cd4_tb__tbscreen_with_demos = pd.merge(hts_art_art_current_status_who_vl_cd4_tb_with_demos, art_current_status_who_vl_cd4_tb_tbscreen_results, on='person_id', how='left')

                        # SQL query to select the specified details for the person_ids
                        if len(hts_art_art_current_status_who_vl_cd4_tb__tbscreen_with_demos["patient_id"]) > 2:
                            art_visit_query = f"""
                            SELECT patient_id, time as art_visit_date
                            FROM patient WHERE patient_id IN {tuple(hts_art_art_current_status_who_vl_cd4_tb__tbscreen_with_demos["patient_id"])}
                            """
             
                            engine = get_db_connection("consultation")
                            art_current_status_who_vl_cd4_tb_tbscreen_art_visit_results = pd.read_sql(art_visit_query, engine)
                            engine.dispose()
                        elif len(hts_art_art_current_status_who_vl_cd4_tb__tbscreen_with_demos["patient_id"]) >= 1 and  len(hts_art_art_current_status_who_vl_cd4_tb__tbscreen_with_demos["patient_id"]) <= 2:
                            person_ids_to_query = ', '.join(f"'{uuid}'" for uuid in hts_art_art_current_status_who_vl_cd4_tb__tbscreen_with_demos["patient_id"])
                            art_visit_query = f"""
                            SELECT patient_id, time as art_visit_date
                            FROM patient WHERE patient_id IN {str(tuple(hts_art_art_current_status_who_vl_cd4_tb__tbscreen_with_demos["patient_id"])).replace(",","")}
                            """
                            # Execute the query and fetch the result into a DataFrame
                            engine = get_db_connection("consultation")
                            art_current_status_who_vl_cd4_tb_tbscreen_art_visit_results = pd.read_sql(art_visit_query, engine)
                            engine.dispose()

                        hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos = pd.merge(hts_art_art_current_status_who_vl_cd4_tb__tbscreen_with_demos, art_current_status_who_vl_cd4_tb_tbscreen_art_visit_results, on='patient_id', how='left')
                        # Replace b'\x00' with 'NO' in the entire DataFrame
                        hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos.drop_duplicates(subset=["person_id", "art_id", "art_visit_date"], inplace=True)
                        hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos = calculate_last_visit_date(hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos)
                        hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos["province"] = province
                        hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos["district"] = district
                        hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos["facility"] = facility
                        hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos["site_id"] = site_id


                    
                        sexual_history_query = """
                                            select sexual_history_id, person_id, sexually_active, date AS sexual_history_date from sexual_history
                                        """
                        engine = get_db_connection("consultation")
                        df_sexual_history = pd.read_sql(sexual_history_query, engine)
                        engine.dispose()



                        # ---------------a3. new HIV positives
                        df_sexual_history_question = """
                                            select sexual_history_id,question,response_type from sexual_history_question 
                                        """
                        
                        engine = get_db_connection("consultation")
                        df_sexual_history_question = pd.read_sql(df_sexual_history_question, engine)
                        engine.dispose()
                        
                        df_sexual_history_question.drop_duplicates(subset=["sexual_history_id","question"],inplace = True)
                        df_sexual_history_question = df_sexual_history_question.pivot(index='sexual_history_id', columns='question')
                        df_sexual_history_question.columns = df_sexual_history_question.columns.droplevel()
                        df_sexual_history_question.reset_index(inplace=True)
                        df_sexual_history_question = pd.merge(df_sexual_history_question, df_sexual_history , how="left", on=["sexual_history_id"])
                        df_sexual_history_question.drop(['sexual_history_id'], axis=1,inplace=True)
                        df_sexual_history_question.rename(columns = {'Exchanged sex for  money/material goods':'Exchanged sex for moneymaterial goods',
                                                                        'Victim/ Suspected victim of sexual abuse':'Victim Suspected victim of sexual abuse',
                                                                        'Unprotected sex without a condom':'Unprotected sex without a condom',
                                                                        'sexually_active': 'Sexually Active' }, inplace = True)
                        
                        hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos = pd.merge(hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos, df_sexual_history_question, on='person_id', how='left')
                        hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos = hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos.applymap(lambda x: "No" if x == b'\x00' else x)
                        hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos = hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos.applymap(lambda x: "Yes" if x == b'\x01' else x)

                        column_order = ['person_id', 'province', 'district', 'facility', 'site_id',
                                        'birthdate', 'sex', 'self_identified_gender', 'marital',
                                        'education', 'occupation', 'religion', 'nationality',
                                        'laboratory_investigation_id', 'hts_number','date_of_hiv_test',
                                        'reason_for_hiv_testing', 'entry_point',
                                        'pregnant_during_hiv_test', 'client_already_positive',
                                        'client_already_on_art', 'test_method', 'hiv_test_one_result', 'hiv_test_two_result', 'hiv_test_three_result','hiv_final_test_result',
                                        'Been incarcerated into jail',
                                        'Exchanged sex for moneymaterial goods', 'Had Anal Sex', 'Had Oral Sex',
                                        'Had sex with a sex worker', 'Had sex with female', 'Had sex with male',
                                        'History of an STI', 'Injected recreational drugs',
                                        'Received blood transfusions', 'Received medical injections',
                                        'Tattooed with unsterilized instruments',
                                        'Unprotected sex (sex without a condom)', 'Used recreational drugs',
                                        'Victim Suspected victim of sexual abuse', 'Sexually Active',
                                        'date_recency_done', 'recency_result', 'reason_for_not_performing_recency_test',
                                        'art_id', 'art_number',
                                        'art_cohort_number', 'Art_Initiation_Date', 'date_enrolled_into_art',
                                        'art_current_status_date', 'regimen', 'arv_status',
                                        'art_initiation_category', 'art_regimen_id','art_visit_date', 'last_art_visit_date',
                                        'who_stage_date','who_outcome_date', 'who_stage', 'art_outcome',
                                        'date_at_which_viral_load_result_was_issued',
                                        'date_for_which_viral_load_was_taken',
                                        'viral_load_sample_submitted_to_lab', 'was_viral_load_result_issued',
                                        'viral_load_result', 'date_at_which_cd4_sample_was_taken',
                                        'date_at_which_cd4_result_was_issued', 'cd4_sample_submitted_to_lab',
                                        'was_cd4_result_issued', 'cd4_count', 'tb_treatment_start_date',
                                        'type_of_tb', 'tb_disease_site', 'tb_disease_type', 'tb_outcome',
                                        'tb_screened']
                        for column in column_order:
                            if column not in hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos.columns:
                                hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos[column] = ''

                        # Now, reorder the columns according to column_order
                        hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos = hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos[column_order]
                        # hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos['art_visit_date'] = pd.to_datetime(art_current_status_who_vl_cd4_tb_tbscreen_art_visit_results['art_visit_date'])
                        print(hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos.head())
                        # hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos.to_csv("./test.csv", index=False)

                        hts_art_art_current_status_who_vl_cd4_tb__tbscreen_artvisit_with_demos.to_excel(f"./extracted_data/{site_id}.xlsx", index=False)


                        with open("./processed.txt", "a")  as log_file:
                            log_file.write(f'{backup_file}\n')
                        print(f"PROCESSED >>> {len(extracted_sites) + 1} <<< SITES")
                        # Commit the transaction if everything is successful
                        session.commit()
                    else:
                        print(f"No POSITIVE clients for {backup_file}")
                        with open("./no_positives.txt", "a")  as f:
                            f.write(f'{backup_file}\n')
            elif not backup_file.endswith(".sql"):
                print(f"{backup_file} Not an sql file! -- Skipping")
            else:
                print(f"{backup_file} already processed! -- Skipping")
            
            engine.dispose()
        
        except Exception as e:
            # Rollback the transaction if an error occurs
            session.rollback()
            # If an error occurs, roll back the transaction
            print(f"An error occurred: {e}")
            with open("./logs.txt", "a")  as log_file:
                log_file.write(f'{backup_file}\n')
            # Optionally, re-raise the exception or handle it as needed
            pass
        finally:
            # Close the session and connection
            session.close()
            engine.dispose()
            

                